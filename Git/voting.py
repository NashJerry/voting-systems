#Name: Panashe Mundondo #StudentNum: 201542826
import collections
import openpyxl
from collections import Counter
import copy

workbook = openpyxl.load_workbook('voting.xlsx')
vSheet = workbook.active
dictionary = {}
print("Hello")
def generatePreferences(fileWithData):
    '''
    Loads an excel file with the voting data from agents and their voting habits for the alternatives
    and puts them in order according to the preference of each agent as a dictionary.

    Parametres:
    fileWithData : the excel file with voting details that will be loaded in to be analysed

    Returns:
    dictionary(dict): a dictionary with keys representing each agent and the values being a list of the agent's preference from highest to lowest
    '''
    for row in range(1, fileWithData.max_row + 1):
        agentIDs = []
        weightOfPreference = []
        inOrder = []
        preferenceOrder = []
        #appends the rows and columns of the file's data to  the weightOfPreference list giving each agent a list of values.
        for column in range(1, fileWithData.max_column + 1):
            weightOfPreference.append(fileWithData.cell(row, column).value)
        #enumerate will give the agents an ascending "agent number" starting at 1. this is then appended to the empty list.
        for agent,candidate in enumerate(weightOfPreference, start = 1):
            inOrder.append((agent,candidate))
        #sorts the list according to which candidate each agent prefers the most
        inOrder.sort(key = lambda x: (x[1], x[0]), reverse=True)
        #dictionary is updated so that it contains tuples as ((Agent Number), (List of Alternative Numbers In Order))
        for x in inOrder:
            preferenceOrder.append(x[0])
        dictionary.update({row: preferenceOrder})
        agentIDs.append(list(dictionary.keys()))
    #returns the dictionary as dictionary to be used in the rest of the program
    return dictionary
generatePreferences(vSheet)



#dictatorship
def dictatorship(preferenceProfile, agent):
    '''
    An agent is selected and whoever they have in first is declared the winner. If the agent number inserted does not exist an error is raised

    Parametres:
    preferenceProfile(dict): Should be a dictionary with all the agents and alternatives they selected
    agent (int) :  An int that maps to the agent's number. Whoever this agent is essentially declares the winner
    
    Returns:
    winner (int): the Alternative number who won according to these rules, or a print statement of an error
    '''
    winner = ''    
    try:
        #if the agent number selected exists in the dictionary keys, return their first place as the winner
        if agent in preferenceProfile:
            winner = preferenceProfile[agent][0] 
        #if the agent number selected does not exist, inform the user
        else:
            raise Exception 
    except Exception:    
        print("Sorry, that number does no map to a candidate!")
    return winner
    
def tieBreak(tiedAlternatives, TiebreakMethod, preferences):
    '''
    Used in subsequent voting systems where draws are a possibility.

    Parametres:
    tiedAlternatives (list): a list of alternative numbers who are in the tied list of the function in which tieBreak is called
    TiebreakMethod (str or int): how the winner of the tie will be decided.
    preferences (dict): a dictionary of the voting preferences

    Returns: 
    winner (int): alternative number who won after the tiebreak.  
    
    '''
    #if min is used the winner is the alternative with the lower number
    if TiebreakMethod == "min":
        winner = min(tiedAlternatives)
        return winner
    #if max is sued, the winner is the alternative with the higher number
    elif TiebreakMethod == "max":
        winner = max(tiedAlternatives)
        return winner
    #if the tiebreak method is an int within the options, the winner is whoever this agent has higher from the tied alternatives
    elif isinstance(TiebreakMethod,int):
        try:
            if TiebreakMethod in preferences:
                lst = []
                for i in tiedAlternatives:
                    lst.append(preferences[TiebreakMethod].index(i))
                winner = (preferences[TiebreakMethod][min(lst)]) 
                return winner
            else:
                raise Exception
        #no alternative has this number
        except KeyError:
            print("Input interger doesn't exist")
    
#scoringRule
def scoringRule(preferences, scoreVector, tiebreak):
    '''
    for each agent, the most preferred alternative is assigned the highest score as determined by the input scoring factor. the score then descends with the preference order

    Parametres:
    preferences (dict) : the dictionary with voting preferences
    scoreVectore (list) : the method the user wishes to assign to each candidate. Allows a custom scoring system
    tiebreak (int or str): how the winner is chosen in the event of a draw

    Returns:
    winner(int): the number of the alternative who wins using these rules
    OR
    Exception(str): an error message if the wrong input is inserted
    
    '''
    tied_values = []
    #scoreVector now goes from highest to lowest making assignement of scores more straight forward
    scoreVector.sort(reverse=True) 
    lengthComparison = []
    totalScores = []
    #lenth of the score vector
    m = len(scoreVector)
    incremontor = Counter()

    
    for x in preferences:
        #check if the length of the score vector is not equal to that of the numbers of alternatives. If they are not equal add 1, else add 0. Will be used for error handling.
        if m != len(preferences[x]):
            lengthComparison.append(1)
        else:
            lengthComparison.append(0) #checking if each voters score is the same as no of alts.
    try:
        #so if the length of the score vector is equivalent to all agents's number of votes this code can run
        if sum(lengthComparison) == 0: 
        #for every agent's preferences, assign them scores according to the input score vector. subsequently append to a list
            for x in preferences: 
                agentScores = {}
                for alternativeNo, atlternativeScore in zip(preferences[x], scoreVector):
                    agentScores.update({alternativeNo: atlternativeScore}) 
                totalScores.append(agentScores) 
                for score in totalScores:
                    incremontor.update(score) 
            #total score for each alternative are added
            TotalSum = dict(incremontor)
            #winner has the highest score 
            winner = max(TotalSum, key = TotalSum.get)
            
            #if the length of unique values in total sum isn't equal to the length of all TotalSum, there must be ties
            if len(set(TotalSum.values())) != len(TotalSum.values()):
                x = max(TotalSum.values())
                #the key of each alternative who has the joint maximum votes is added to tied values
                tied_values = [key for key, value in TotalSum.items() if value == x]
            #if the length of tied values is less than 2 there is no draw so a direct winner can be found. otherwise enter the tieBreak function
            if len(tied_values) == 0:
                return winner
            elif winner in tied_values and len(tied_values) == 1:
                return winner
            else:
                return tieBreak(tied_values, tiebreak, preferences)
        #if the sum of the length comparison is not 0, either an agent does not have enough votes or the scorevector is not the right length. return an erros
        else:
            raise Exception
    except Exception:
        print("Incorrect Input")
    return winner

def plurality(preferences, tiebreak):
    '''
    The winner is the agent who appeared in first place most often overall.

    Parametres:
    preferences (dict): the dictionary with the voting details

    Returns:
    winner (int): the winner either directly or after a tiebreak
    
    '''
    firstPick = []
    tied_values = []
    firstPickDictionary = {}
    #create a list with each agent's first place pick
    for i in preferences:
        firstPick.append(preferences[i][0])
    #count how often each agent was in first place. The winner then has the highest score
    firstPickDictionary = dict(collections.Counter(firstPick))
    winner = max(firstPick, key = firstPick.count)
    #if the length of unique values in firstPickDictionary isn't equal to the length of all firstPickDictionary, there must be ties
    if len(set(firstPickDictionary.values())) != len(firstPickDictionary.values()):
        x = max(firstPickDictionary.values())
    #the key of each alternative who has the joint maximum votes is added to tied values
        tied_values = [key for key, value in firstPickDictionary.items() if value == x]
    #if the length of tied values is less than 2 there is no draw so a direct winner can be found. otherwise enter the tieBreak function
    if len(tied_values) == 0:
        return winner
    elif winner in tied_values and len(tied_values) == 1:
        return winner
    else:
        return tieBreak(tied_values, tiebreak, preferences)


def borda(preferences, tiebreak): 
    '''
    If there are m candidates, each at position j, each agent's least preferred candidate receives a 0. 
    their favourite receives a score of m-1, meaning every agent gets a score of m-j. 

    Parametres:
    preferences(dict): A dictionary with the voting details

    Returns:
    winner (int): the winner using this voting system
    
    '''
    tied_values = []
    OverallScore = [] 
    incrementor = Counter() 
    for p in preferences:
        Dict_in_Order1 = {}
        List_in_order1 = []
        #creates a list enumerating all the alternatives from biggest to smallest. the dictionary and overall score are then updated and appended
        for x, y in reversed(list(enumerate(reversed(preferences[p])))):
            List_in_order1.append((y, x))
        Dict_in_Order1.update(List_in_order1)
        OverallScore.append(Dict_in_Order1)
    #Accumulate the scores that each alternative got. The winner is the one with the highest score
    for o in OverallScore:
         incrementor.update(o)
    FinalScore = dict(incrementor)
    winner = max(FinalScore, key = FinalScore.get)
    #if the length of unique values in FinalScore isn't equal to the length of all FinalScore, there must be ties
    if len(set(FinalScore.values())) != len(FinalScore.values()):
        x = max(FinalScore.values())
    #the key of each alternative who has the joint maximum votes is added to tied values
        tied_values = [key for key, value in FinalScore.items() if value == x]
    #if the length of tied values is less than 2 there is no draw so a direct winner can be found. otherwise enter the tieBreak function
    if len(tied_values) == 0:
        return winner
    elif winner in tied_values and len(tied_values) == 1:
        return winner
    else:
        return tieBreak(tied_values, tiebreak, preferences)


#HARMOIC
def harmonic(preferences, tiebreak):
    '''
    An agent at position x receives a score of 1/x. The agent with the highest score wins.

    Parametres:
    preferences (dict): a dictionary with all voting choices
    tiebreak (str or int) : the method that would be used to break any tiebreaks 

    Returns:
    winner (int): the winner using this voting system
    
    '''

    incrementor = Counter() 
    InitialList = []
    tied_values = []
    for p in preferences:
        Dict_in_Order1 = {}
        List_in_order1 = []
        #for each agent's voting list, they are enumerated begininng with one as as default index is 0.
        #each is then assigned a score of 1/their index.
        for x, y in (list(enumerate((preferences[p]), start = 1))):
            List_in_order1.append((y, 1/x))
        Dict_in_Order1.update(List_in_order1)
        InitialList.append(Dict_in_Order1)
    for n in InitialList:
        incrementor.update(n)
    #counter is used to add up the final score
    BFinalScore = dict(incrementor)
    winner = max(BFinalScore, key = BFinalScore.get) 
    #if the length of unique values in FinalScore isn't equal to the length of all FinalScore, there must be ties
    if len(set(BFinalScore.values())) != len(BFinalScore.values()):
        x = max(BFinalScore.values())
        #the key of each alternative who has the joint maximum votes is added to tied values
        tied_values = [key for key, value in BFinalScore.items() if value == x]
    #if the length of tied values is less than 2 there is no draw so a direct winner can be found. otherwise enter the tieBreak function
    if len(tied_values) == 0:
        return winner
    elif winner in tied_values and len(tied_values) == 1:
        return winner
    else:
        return tieBreak(tied_values, tiebreak, preferences)


#VETO
def veto(preferences, tiebreak):
    '''
    The alternative in each agent's last position gets 0 points, all others get 1. 
    The winner is the alternative with the most points

    Parametres:
    preferences (dict): a dictionary with all voting choices
    tiebreak (str or int) : the method that would be used to break any tiebreaks 

    Returns:
    winner (int): the winner using this voting system
    
    '''
    scoringList = []
    incrementor = Counter()
    tied_values = []
    for p in preferences:
    #in each agent's voting list, if the scoring list is shorter than the voting list - 1, 1 is appended
        if len(scoringList) < len(preferences[p]) -1:
            scoringList.append(1)
    #otherwise 0 is appended. This obtains a list of all 1s with the last element being 0
    scoringList.append(0)
    OverallScore = []
    for i in preferences:
        emptyDic ={}
        #creares a dictionary with each agent being awared a score corresponding to the scoringList. This is appended to a list
        for a,b in zip(preferences[i], scoringList):
            emptyDic.update({a:b})
        OverallScore.append(emptyDic)

#the score for each alternative is summed  
    for o in OverallScore: 
        incrementor.update(o)
    FinalScore = dict(incrementor)
    winner = max(FinalScore, key = FinalScore.get)

    #if the length of unique values in FinalScore isn't equal to the length of all FinalScore, there must be ties
    if len(set(FinalScore.values())) != len(FinalScore.values()):
        x = max(FinalScore.values())
    #the key of each alternative who has the joint maximum votes is added to tied values
        tied_values = [key for key, value in FinalScore.items() if value == x]

    #if the length of tied values is less than 2 there is no draw so a direct winner can be found. otherwise enter the tieBreak function
    if len(tied_values) == 0:
        return winner
    elif winner in tied_values and len(tied_values) == 1:   
        return winner
    else:
        return tieBreak(tied_values, tiebreak, preferences)

def rangeVoting(values, tiebreak): 
    '''
    Adds each alternatives valuations from an excel sheet

    Parametres:
    values (xlsx): a worksheet with voting valuations
    tiebreak : a tiebreak method

    Returns:
    winner (int): the winner using this voting system
    
    '''

    ovrScore = []
    incrementor = Counter()
    tied_values = []
    EmptyDic ={}
    #use data within these rows of the excel file
    for row in range(1, values.max_row + 1):
        agentIDs = []
        weightOfPreference = []
        inOrder = []
        preferenceOrder = []
        #appends the numerical valuation of each agent on each candidate
        for column in range(1, values.max_column + 1):
            weightOfPreference.append(values.cell(row, column).value)
        #candidates are enumerated starting from one and mapped onto a weight of preference value. Appended to the InOrder list
        for agent,candidate in enumerate(weightOfPreference, start = 1):
            inOrder.append((agent,candidate))
        #sorted to have the most preferred candidate first by using the reverse (which was naturally ascending before)
        inOrder.sort(key = lambda x: (x[1], x[0]), reverse=True)
        #preferenceOrder is updated with just the alternative numbers
        for x in inOrder:
            preferenceOrder.append(x[0])
        #dictionary updated to account for the preference order of each candidate
        dictionary.update({row: preferenceOrder})
        agentIDs.append(list(dictionary.keys()))    
        #Dictionary is updated to show the values associated with each agent's voting in order  
        EmptyDic.update(inOrder)
        ovrScore.append(EmptyDic)

    #the total score for each alternative is calculated. Winner declared
    for o in ovrScore:
        incrementor.update(o)
    FinalScore = dict(incrementor)
    winner = max(FinalScore, key = FinalScore.get)

    #if the length of unique values in FinalScore isn't equal to the length of all FinalScore, there must be ties
    if len(set(FinalScore.values())) != len(FinalScore.values()):
        x = max(FinalScore.values())
    #the key of each alternative who has the joint maximum votes is added to tied values
        tied_values = [key for key, value in FinalScore.items() if value == x]
    #if the length of tied values is less than 2 there is no draw so a direct winner can be found. otherwise enter the tieBreak function    
    if len(tied_values) == 0:
        return winner
    elif winner in tied_values and len(tied_values) == 1:       
        return winner
    else:
        return tieBreak(tied_values, tiebreak, dictionary)


def STV(preferences, tiebreak):
    #create a deep copy of the dictionary in order not to alter the actual values
    DicCopy = copy.deepcopy(preferences) 
    frequency = {}

    '''
    knocks out the lowest voted candidate each round unitl one winner remains

    Parametres:
    preferences (dict): a worksheet with voting valuations
    tiebreak (int or string) : a tiebreak method

    Returns:
    winner (int): the winner using this voting system
    
    '''
    for i in DicCopy:
        firstPicks = []
        #Ensures this runs until there is one possible winner left
        while len(DicCopy[i])>1:
            for i in DicCopy:
                #appends and counts the first place position of each agent
                firstPicks.append(DicCopy[i][0])
                frequency = dict(collections.Counter(firstPicks))

                #assigns 0 to any alternative who did not appear in first even once
                for x in DicCopy[i]:
                    if x not in firstPicks:
                        frequency.update({x:0})
                #assigns the agent who was in first the least
                minimum = min(frequency, key = frequency.get)
            
            #the minimum value is removed from the copy of the dictionary. First pick is cleared for the beginning of the 
            #next loop. minimum also removed from the prequency
            for i in DicCopy:  
                DicCopy[i].remove(minimum)
                firstPicks.clear()
            frequency.pop(minimum)
    #If there is more than one remaining candidate, there is a tie and tiebreak is called
    if len(frequency) > 1:
        return tieBreak(frequency, tiebreak, dictionary)
    else:
        return DicCopy[i][0]